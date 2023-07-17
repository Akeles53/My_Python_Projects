from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import *
from tkinter import messagebox
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import time


options = Options()
options.headless = False

icons = {"adress": "place_gm_blue_24dp.png",
         "phone": "phone_gm_blue_24dp.png",
         "wsite": "public_gm_blue_24dp.png",
         "points": "ic_plus_code.png"}

def main(input_sector,input_district,input_city,input_country,input_count,save):
    global icons
    class musteri():
        def __init__(self, name, score, pNumber, adress, sector, site):
            self.name = name
            self.score = score
            self.pNumber = pNumber
            self.adress = adress
            self.sector = sector
            self.site = site

        def info(self):
            print("----------------------------------")
            print("Name: "+self.name,"\nScore: "+ self.score,"\nPhone Number: "+ self.pNumber,"\nAdress: "+ self.adress,"\nSector: "+self.sector,"\nWeb Site: "+ self.site)
            print("----------------------------------")
    def write_info(sira,save):
        wb = load_workbook("./excel_files/"+str(save) + ".xlsx")
        ws = wb.active
        sira += 1
        ws['A' + str(sira)].value = musteriler[-1].name
        ws['B' + str(sira)].value = musteriler[-1].score
        ws['C' + str(sira)].value = musteriler[-1].adress
        ws['D' + str(sira)].value = musteriler[-1].pNumber
        ws['E' + str(sira)].value = musteriler[-1].sector
        ws['F' + str(sira)].value = musteriler[-1].site
        wb.save("./excel_files/"+str(save) + ".xlsx")

    def scroll(x):
        global els
        while True:
            els = driver.find_elements(By.CLASS_NAME, 'TFQHme')
            print(len(els))
            driver.execute_script("arguments[0].scrollIntoView();", els[-1])
            try:
                driver.execute_script("arguments[0].scrollIntoView();", els[1])
            except:
                htmlelement = driver.find_element(By.TAG_NAME, 'html')
                htmlelement.send_keys(Keys.HOME)

            time.sleep(0.5)
            if len(els) >= x + 5:
                print("Yeterli eleman bulundu")
                break
            try:
                kd = driver.find_element(By.CLASS_NAME,"HlvSq")
                print("Tüm elemanlar bulundu")
                break
            except:
                pass
    def get_info():
        isim = "----"
        score = "----"
        sector = "----"
        adress = "----"
        wsite = "----"
        phone = "----"
        try:
            time.sleep(1)
            isim = driver.find_element(By.XPATH, "//div[@class='lMbq3e']/div/h1").text

            if musteriler[-1].name == isim and len(musteriler) > 0:
                for z in range(0,8):
                    time.sleep(0.5)
                    isim = driver.find_element(By.XPATH, "//div[@class='lMbq3e']/div/h1").text
                    if isim != musteriler[-1].name:
                        break

            score = driver.find_element(By.XPATH,
                                        "/html/body/div[3]/div[9]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div[1]/div[1]/div[2]/div/div[1]/div[2]/span[1]/span/span[1]").text
            sector = driver.find_element(By.XPATH,
                                         "/html/body/div[3]/div[9]/div[9]/div/div/div[1]/div[3]/div/div[1]/div/div/div[2]/div[2]/div[1]/div[1]/div[2]/div/div[2]/span[1]/span[1]/button").text

        except:
            if isim == "----":
                time.sleep(2)
                isim = driver.find_element(By.XPATH, "//div[@class='lMbq3e']/div/h1").text
        text = "----"
        srcc = "----"
        srcs = driver.find_elements(By.CLASS_NAME, "Liguzb")
        texts = driver.find_elements(By.CLASS_NAME, "rogA2c")
        w_sites = driver.find_elements(By.CLASS_NAME, "CsEnBe")
        w_site_array = []
        for i in range(0,len(srcs)):
            srcc = srcs[i].get_attribute("src")
            srcc_sp = srcc.split("/")
            text = texts[i].text
            if srcc_sp[-1] == icons["adress"]:
                adress = text
            if srcc_sp[-1] == icons["wsite"]:
                for w in w_sites:
                    try:
                        if w.get_attribute("href") != None:
                            w_site_array.append(w.get_attribute("href"))

                    except:
                        print("denendi")
                try:
                    wsite = w_site_array[-1]
                except:
                    pass
            if srcc_sp[-1] == icons["phone"]:
                phone = text
        try:
            wsite = w_site_array[-1]
        except:
            pass
        musteriler.append(musteri(isim, score, phone, adress, sector, wsite))

    def creating_excel(save):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        ws['A1'].value = "Name"
        ws['B1'].value = "Score"
        ws['C1'].value = "Adress"
        ws['D1'].value = "Phone Number"
        ws['E1'].value = "Sector"
        ws['F1'].value = "Web Site"

        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 6
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 17
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 34

        wb.save("./excel_files/"+save+".xlsx")

    creating_excel(save)
    musteriler = []
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get("https://www.google.com/maps/")

    time.sleep(3)
    search_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#searchboxinput")))
    search_input.send_keys("{}/{}/{}".format(input_district, input_city, input_country))
    search_button= driver.find_element(By.CSS_SELECTOR, "#searchbox-searchbutton").click()
    time.sleep(4)

    try:
        clear_button = driver.find_element(By.XPATH,"/html/body/div[3]/div[9]/div[3]/div[1]/div[1]/div[1]/div[2]/a/span").click()
    except:
        time.sleep(1)
        clear_button = driver.find_element(By.XPATH,"/html/body/div[3]/div[9]/div[3]/div[1]/div[1]/div/div[2]/div[3]/button").click()

    time.sleep(3)

    search_input.send_keys("{} near {}/{}/{}".format(input_sector, input_district, input_city, input_country))

    driver.find_element(By.XPATH,"/html/body/div[3]/div[9]/div[3]/div[1]/div[1]/div/div[2]/div[1]/button").click()
    time.sleep(6)

    scroll(input_count)


    sectors = driver.find_elements(By.CLASS_NAME,"hfpxzc")
    print(len(sectors))
    sira = 0

    for sector in sectors:
        sira = sira + 1

        try:
            sector.click()
            get_info()
            musteriler[-1].info()
        except:
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", sector)
            get_info()
        write_info(sira,save)

    driver.quit()

window = Tk()
window.title("g_maps")
window.geometry("300x400")

L1 = Label(window,text="Ülke:")
L1.place(relx=0.1,rely=0.1)
E1 = Entry(window,bd=5)
E1.insert(END,"Türkiye")
E1.place(relx=0.5,rely=0.1)
L2 = Label(window,text="Şehir:")
L2.place(relx=0.1,rely=0.2)
E2 = Entry(window,bd=5)
E2.insert(END,"İstanbul")
E2.place(relx=0.5,rely=0.2)
L3 = Label(window,text="İlçe:")
L3.place(relx=0.1,rely=0.3)
E3 = Entry(window,bd=5)
E3.place(relx=0.5,rely=0.3)
L4 = Label(window,text="Anahtar Kelime:")
L4.place(relx=0.1,rely=0.4)
E4 = Entry(window,bd=5)
E4.place(relx=0.5,rely=0.4)
L5 = Label(window,text="Adet:")
L5.place(relx=0.1,rely=0.5)
E5 = Entry(window,bd=5)
E5.place(relx=0.5,rely=0.5)
L6 = Label(window,text="Dosya Adı:")
L6.place(relx=0.1,rely=0.6)
E6 = Entry(window,bd=5)
E6.place(relx=0.5,rely=0.6)
L7 = Label(window,text="Not: Kelimeleri doğru yazmaya özen gösterin.")
L7.place(relx=0.1,rely=0.9)

def ButtonFunc():
    if len(E1.get()) < 1:
        messagebox.showinfo("Bilgilendirme","Ülke boş olmamalı.")
    elif len(E2.get()) < 1:
        messagebox.showinfo("Bilgilendirme","Şehir boş olmamalı.")
    elif len(E3.get()) < 1:
        messagebox.showinfo("Bilgilendirme", "İlçe boş olmamalı.")
    elif len(E4.get()) < 1:
        messagebox.showinfo("Bilgilendirme", "Anahtar kelime boş olmamalı.")
    elif len(E5.get()) < 1:
        messagebox.showinfo("Bilgilendirme", "Adet boş olmamalı.")
    elif len(E6.get()) < 1:
        messagebox.showinfo("Bilgilendirme", "Dosya adı boş olmamalı.")
    else:
        try:
            if int(E5.get()) < 1:
                messagebox.showinfo("Bilgilendirme","Adet 0'dan büyük olmalıdır.")
            else:
                try:
                    sector = str(E4.get())
                    district = str(E3.get())
                    city = str(E2.get())
                    country = str(E1.get())
                    count = int(E5.get())
                    save = str(E6.get())
                    main(sector, district, city, country, count, save)
                    messagebox.showinfo("Bilgilendirme","Program Başarıyla Tamamlandı")
                except:
                    messagebox.showerror("Hata", "Bir sorun oldu.")
        except ValueError:
            messagebox.showinfo("Bilgilendirme","Adet sayı olmalı.")



B = Button(window, text ="Başlat", command = ButtonFunc)
B.place(relx=0.5,rely=0.8,anchor=CENTER)

window.mainloop()